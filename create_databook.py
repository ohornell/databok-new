"""
Skapa Freemelt finansiell databok för Q1-Q3 2025
Investment Bank formatering (Goldman Sachs-stil)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================
# INVESTMENT BANK STYLE GUIDE
# ============================================
# Typsnitt: Arial eller Calibri, 8-10pt
# Färger: Minimalistiskt - svart, grått, mörkblått
# Linjer: Tunna linjer, dubbla linjer för totaler
# Negativa tal: Parenteser, inte minustecken
# Tusentalsavgränsare: Komma
# Decimaler: Inga för heltal, en decimal för %

# Färgpalett (Goldman Sachs-inspirerad)
GS_NAVY = "1F3864"  # Mörkblå för headers
GS_LIGHT_BLUE = "D6DCE4"  # Ljusblå för subtotaler
GS_LIGHT_GRAY = "F2F2F2"  # Ljusgrå för totaler
GS_DARK_GRAY = "404040"  # Mörkgrå för text
GS_BLACK = "000000"

# Färgkodning för data (xlsx skill standard)
COLOR_HARDCODED = "0000FF"  # Blå - hårdkodade värden/inputs
COLOR_FORMULA = "000000"    # Svart - formler
COLOR_LINK = "008000"       # Grön - cross-sheet länkar

# Typsnitt
TITLE_FONT = Font(name='Arial', size=11, bold=True, color=GS_NAVY)
HEADER_FONT = Font(name='Arial', size=9, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name='Arial', size=8, italic=True, color=GS_DARK_GRAY)
SECTION_FONT = Font(name='Arial', size=9, bold=True, color=GS_NAVY)
LABEL_FONT = Font(name='Arial', size=9, color=GS_DARK_GRAY)
DATA_FONT_HARDCODED = Font(name='Arial', size=9, color=COLOR_HARDCODED)  # Blå för inputs
DATA_FONT_FORMULA = Font(name='Arial', size=9, color=COLOR_FORMULA)      # Svart för formler
DATA_FONT_LINK = Font(name='Arial', size=9, color=COLOR_LINK)            # Grön för länkar
TOTAL_FONT = Font(name='Arial', size=9, bold=True, color=GS_BLACK)
TOTAL_FONT_FORMULA = Font(name='Arial', size=9, bold=True, color=COLOR_FORMULA)
SUBTOTAL_FONT = Font(name='Arial', size=9, bold=True, color=GS_DARK_GRAY)
SUBTOTAL_FONT_FORMULA = Font(name='Arial', size=9, bold=True, color=COLOR_FORMULA)
SOURCE_FONT = Font(name='Arial', size=7, italic=True, color="808080")

# Fyllningar
HEADER_FILL = PatternFill(start_color=GS_NAVY, end_color=GS_NAVY, fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color=GS_LIGHT_BLUE, end_color=GS_LIGHT_BLUE, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=GS_LIGHT_GRAY, end_color=GS_LIGHT_GRAY, fill_type="solid")

# Ramar
thin_side = Side(style='thin', color=GS_DARK_GRAY)
hair_side = Side(style='hair', color="BFBFBF")
double_side = Side(style='double', color=GS_BLACK)
medium_side = Side(style='medium', color=GS_DARK_GRAY)

HEADER_BORDER = Border(bottom=medium_side)
NO_BORDER = Border()  # Ingen kantlinje för vanliga datarader
SUBTOTAL_BORDER = Border(top=thin_side, bottom=thin_side)
TOTAL_BORDER = Border(top=thin_side, bottom=double_side)
SECTION_BORDER = Border(bottom=thin_side)

# Alignment
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', indent=1)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
INDENT_ALIGN = Alignment(horizontal='left', vertical='center', indent=2)

# Nummerformat (parenteser för negativa)
NUMBER_FORMAT = '#,##0_);(#,##0);"-"_)'
PERCENT_FORMAT = '0.0%_);(0.0%)'


def apply_header_row(ws, row, headers):
    """Applicera header-formatering - första kolumnen vänster, resten höger"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN  # Kvartal högerjusterade
        cell.border = HEADER_BORDER


def apply_data_row(ws, row, values, row_type='data'):
    """
    Applicera datarad med korrekt formatering
    row_type: 'data', 'subtotal', 'total', 'section', 'blank'
    Färgkodning: Blå=hårdkodad data (inkl 0), Svart=formler, Grön=cross-sheet länkar
    """
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)

        # Bestäm datatyp för färgkodning
        is_formula = isinstance(value, str) and value.startswith("=")
        is_link = is_formula and ("!" in value)  # Cross-sheet länk
        # Hårdkodad data inkluderar 0 (int eller float i datakolumner)
        is_hardcoded = isinstance(value, (int, float)) and col > 1

        if row_type == 'section':
            cell.font = SECTION_FONT
            cell.alignment = LEFT_ALIGN
            # Sektionslinje på alla kolumner
            cell.border = SECTION_BORDER
        elif row_type == 'subtotal':
            cell.fill = SUBTOTAL_FILL
            cell.border = SUBTOTAL_BORDER
            cell.alignment = RIGHT_ALIGN if col > 1 else LEFT_ALIGN
            # Färgkodning för subtotaler (inkluderar 0-värden)
            if col > 1 and (value is not None and value != ""):
                if is_link:
                    cell.font = Font(name='Arial', size=9, bold=True, color=COLOR_LINK)
                elif is_formula:
                    cell.font = SUBTOTAL_FONT_FORMULA
                elif is_hardcoded:
                    cell.font = Font(name='Arial', size=9, bold=True, color=COLOR_HARDCODED)
                else:
                    cell.font = SUBTOTAL_FONT
                cell.number_format = NUMBER_FORMAT
            else:
                cell.font = SUBTOTAL_FONT
        elif row_type == 'total':
            cell.fill = TOTAL_FILL
            cell.border = TOTAL_BORDER
            cell.alignment = RIGHT_ALIGN if col > 1 else LEFT_ALIGN
            # Färgkodning för totaler (inkluderar 0-värden)
            if col > 1 and (value is not None and value != ""):
                if is_link:
                    cell.font = Font(name='Arial', size=9, bold=True, color=COLOR_LINK)
                elif is_formula:
                    cell.font = TOTAL_FONT_FORMULA
                elif is_hardcoded:
                    cell.font = Font(name='Arial', size=9, bold=True, color=COLOR_HARDCODED)
                else:
                    cell.font = TOTAL_FONT
                cell.number_format = NUMBER_FORMAT
            else:
                cell.font = TOTAL_FONT
        elif row_type == 'blank':
            pass
        else:  # data
            cell.alignment = RIGHT_ALIGN if col > 1 else INDENT_ALIGN
            cell.border = NO_BORDER
            if col > 1 and (value is not None and value != ""):
                # Färgkodning baserat på datatyp (inkluderar 0-värden)
                if is_link:
                    cell.font = DATA_FONT_LINK
                elif is_formula:
                    cell.font = DATA_FONT_FORMULA
                elif is_hardcoded:
                    cell.font = DATA_FONT_HARDCODED
                cell.number_format = NUMBER_FORMAT
            elif col == 1:
                cell.font = LABEL_FONT


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # ============================================
    # FLIK 1: RESULTATRÄKNING (Income Statement)
    # ============================================
    ws = wb.create_sheet("Income Statement")

    # Titel
    ws.merge_cells('A1:D1')
    ws['A1'] = "FREEMELT HOLDING AB (PUBL)"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN

    ws.merge_cells('A2:D2')
    ws['A2'] = "Consolidated Income Statement"
    ws['A2'].font = Font(name='Arial', size=10, color=GS_DARK_GRAY)

    # Headers med valuta i första kolumnen
    apply_header_row(ws, 4, ["SEK '000", "Q1 2025", "Q2 2025", "Q3 2025"])

    # Period subheader
    ws['A5'] = ""
    ws['B5'] = "Jan-Mar"
    ws['C5'] = "Apr-Jun"
    ws['D5'] = "Jul-Sep"
    for col in range(1, 5):
        ws.cell(row=5, column=col).font = SUBHEADER_FONT
        ws.cell(row=5, column=col).alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN  # Högerjusterade

    # Data
    row = 7

    # Revenue section (row 7)
    apply_data_row(ws, row, ["Revenue", "", "", ""], 'section')
    row += 1  # row 8
    apply_data_row(ws, row, ["Net sales", 2926, 19074, 16988])
    row += 1  # row 9
    apply_data_row(ws, row, ["Capitalized development costs", 3844, 5660, 2998])
    row += 1  # row 10
    apply_data_row(ws, row, ["Other operating income", 1100, 1558, 1403])
    row += 1  # row 11
    apply_data_row(ws, row, ["Total revenue", "=SUM(B8:B10)", "=SUM(C8:C10)", "=SUM(D8:D10)"], 'subtotal')
    row += 2  # row 13

    # Operating expenses section
    apply_data_row(ws, row, ["Operating expenses", "", "", ""], 'section')
    row += 1  # row 14
    apply_data_row(ws, row, ["Cost of goods sold", -529, -10415, -8430])
    row += 1  # row 15
    apply_data_row(ws, row, ["Other external costs", -6827, -9383, -7601])
    row += 1  # row 16
    apply_data_row(ws, row, ["Personnel costs", -10226, -11029, -11066])
    row += 1  # row 17
    apply_data_row(ws, row, ["Depreciation & amortization", -13856, -14955, -14910])
    row += 1  # row 18
    apply_data_row(ws, row, ["Other operating expenses", -516, -1038, -734])
    row += 1  # row 19
    apply_data_row(ws, row, ["Total operating expenses", "=SUM(B14:B18)", "=SUM(C14:C18)", "=SUM(D14:D18)"], 'subtotal')
    row += 2  # row 21

    # Operating result
    apply_data_row(ws, row, ["Operating income (EBIT)", "=B11+B19", "=C11+C19", "=D11+D19"], 'total')
    row += 2  # row 23

    # Financial items
    apply_data_row(ws, row, ["Financial items", "", "", ""], 'section')
    row += 1  # row 24
    apply_data_row(ws, row, ["Interest income", 136, 328, 183])
    row += 1  # row 25
    apply_data_row(ws, row, ["Interest expense", -108, -10, -2])
    row += 1  # row 26
    apply_data_row(ws, row, ["Net financial items", "=B24+B25", "=C24+C25", "=D24+D25"], 'subtotal')
    row += 2  # row 28

    # Pre-tax and net income
    apply_data_row(ws, row, ["Income before tax (EBT)", "=B21+B26", "=C21+C26", "=D21+D26"], 'subtotal')
    row += 1  # row 29
    apply_data_row(ws, row, ["Income tax", 0, -2, 0])
    row += 2  # row 31
    apply_data_row(ws, row, ["Net income", "=B28+B29", "=C28+C29", "=D28+D29"], 'total')
    row += 3

    # Source
    ws.cell(row=row, column=1, value="Source: Freemelt Interim Reports Q1-Q3 2025, page 18").font = SOURCE_FONT

    # Kolumnbredder
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # Frys header
    ws.freeze_panes = 'A6'

    # Ta bort stödlinjer (gridlines)
    ws.sheet_view.showGridLines = False

    # ============================================
    # FLIK 2: BALANSRÄKNING (Balance Sheet)
    # ============================================
    ws = wb.create_sheet("Balance Sheet")

    # Titel
    ws.merge_cells('A1:D1')
    ws['A1'] = "FREEMELT HOLDING AB (PUBL)"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN

    ws.merge_cells('A2:D2')
    ws['A2'] = "Consolidated Balance Sheet"
    ws['A2'].font = Font(name='Arial', size=10, color=GS_DARK_GRAY)

    # Headers med valuta i första kolumnen
    apply_header_row(ws, 4, ["SEK '000", "Q1 2025", "Q2 2025", "Q3 2025"])

    # Period subheader (som text för att undvika Excel-varning)
    ws['A5'] = ""
    ws['B5'] = "31-Mar-2025"
    ws['C5'] = "30-Jun-2025"
    ws['D5'] = "30-Sep-2025"
    for col in range(1, 5):
        cell = ws.cell(row=5, column=col)
        cell.font = SUBHEADER_FONT
        cell.alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN  # Högerjusterade
        if col > 1:
            cell.number_format = '@'  # Textformat för att undvika datumtolkning

    row = 7

    # ASSETS (row 7)
    apply_data_row(ws, row, ["ASSETS", "", "", ""], 'section')
    row += 2  # row 9

    # Non-current assets
    apply_data_row(ws, row, ["Non-current assets", "", "", ""], 'section')
    row += 1  # row 10
    apply_data_row(ws, row, ["Goodwill", 70163, 58282, 46401])
    row += 1  # row 11
    apply_data_row(ws, row, ["Capitalized development", 87777, 91121, 91804])
    row += 1  # row 12
    apply_data_row(ws, row, ["Patents", 3997, 4613, 5214])
    row += 1  # row 13
    apply_data_row(ws, row, ["Total intangible assets", "=SUM(B10:B12)", "=SUM(C10:C12)", "=SUM(D10:D12)"], 'subtotal')
    row += 1  # row 14
    apply_data_row(ws, row, ["Machinery & equipment", 8869, 5165, 9937])
    row += 1  # row 15
    apply_data_row(ws, row, ["Fixtures & fittings", 1112, 1292, 1170])
    row += 1  # row 16
    apply_data_row(ws, row, ["Total tangible assets", "=SUM(B14:B15)", "=SUM(C14:C15)", "=SUM(D14:D15)"], 'subtotal')
    row += 1  # row 17
    apply_data_row(ws, row, ["Deferred tax assets", 5230, 5230, 5230])
    row += 1  # row 18
    apply_data_row(ws, row, ["Total non-current assets", "=B13+B16+B17", "=C13+C16+C17", "=D13+D16+D17"], 'subtotal')
    row += 2  # row 20

    # Current assets
    apply_data_row(ws, row, ["Current assets", "", "", ""], 'section')
    row += 1  # row 21
    apply_data_row(ws, row, ["Inventory", 17146, 20906, 18719])
    row += 1  # row 22
    apply_data_row(ws, row, ["Trade receivables", 11573, 19829, 7646])
    row += 1  # row 23
    apply_data_row(ws, row, ["Other receivables", 1995, 2279, 1774])
    row += 1  # row 24
    apply_data_row(ws, row, ["Prepaid expenses", 3460, 3032, 3040])
    row += 1  # row 25
    apply_data_row(ws, row, ["Cash and cash equivalents", 71822, 53984, 45081])
    row += 1  # row 26
    apply_data_row(ws, row, ["Total current assets", "=SUM(B21:B25)", "=SUM(C21:C25)", "=SUM(D21:D25)"], 'subtotal')
    row += 2  # row 28

    apply_data_row(ws, row, ["TOTAL ASSETS", "=B18+B26", "=C18+C26", "=D18+D26"], 'total')
    row += 3

    # EQUITY AND LIABILITIES (row 31)
    apply_data_row(ws, row, ["EQUITY AND LIABILITIES", "", "", ""], 'section')
    row += 2  # row 33

    # Equity
    apply_data_row(ws, row, ["Shareholders' equity", "", "", ""], 'section')
    row += 1  # row 34
    apply_data_row(ws, row, ["Share capital", 9438, 9438, 9438])
    row += 1  # row 35
    apply_data_row(ws, row, ["Other contributed capital", 534090, 533830, 533830])
    row += 1  # row 36
    apply_data_row(ws, row, ["Retained earnings", -287980, -307781, -328724])
    row += 1  # row 37
    apply_data_row(ws, row, ["Total shareholders' equity", "=SUM(B34:B36)", "=SUM(C34:C36)", "=SUM(D34:D36)"], 'subtotal')
    row += 2  # row 39

    # Current liabilities
    apply_data_row(ws, row, ["Current liabilities", "", "", ""], 'section')
    row += 1  # row 40
    apply_data_row(ws, row, ["Trade payables", 3155, 8429, 9233])
    row += 1  # row 41
    apply_data_row(ws, row, ["Other liabilities", 5583, 4028, 3044])
    row += 1  # row 42
    apply_data_row(ws, row, ["Accrued expenses", 18742, 16671, 9155])
    row += 1  # row 43
    apply_data_row(ws, row, ["Total current liabilities", "=SUM(B40:B42)", "=SUM(C40:C42)", "=SUM(D40:D42)"], 'subtotal')
    row += 2  # row 45

    apply_data_row(ws, row, ["TOTAL EQUITY AND LIABILITIES", "=B37+B43", "=C37+C43", "=D37+D43"], 'total')
    row += 3

    # Source
    ws.cell(row=row, column=1, value="Source: Freemelt Interim Reports Q1-Q3 2025, page 19").font = SOURCE_FONT

    # Kolumnbredder
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    ws.freeze_panes = 'A6'

    # Ta bort stödlinjer (gridlines)
    ws.sheet_view.showGridLines = False

    # ============================================
    # FLIK 3: KASSAFLÖDESANALYS (Cash Flow)
    # ============================================
    ws = wb.create_sheet("Cash Flow")

    # Titel
    ws.merge_cells('A1:D1')
    ws['A1'] = "FREEMELT HOLDING AB (PUBL)"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN

    ws.merge_cells('A2:D2')
    ws['A2'] = "Consolidated Statement of Cash Flows"
    ws['A2'].font = Font(name='Arial', size=10, color=GS_DARK_GRAY)

    # Headers med valuta i första kolumnen
    apply_header_row(ws, 4, ["SEK '000", "Q1 2025", "Q2 2025", "Q3 2025"])

    # Period subheader
    ws['A5'] = ""
    ws['B5'] = "Jan-Mar"
    ws['C5'] = "Apr-Jun"
    ws['D5'] = "Jul-Sep"
    for col in range(1, 5):
        ws.cell(row=5, column=col).font = SUBHEADER_FONT
        ws.cell(row=5, column=col).alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN  # Högerjusterade

    row = 7

    # Operating activities (row 7)
    apply_data_row(ws, row, ["Cash flows from operating activities", "", "", ""], 'section')
    row += 1  # row 8
    apply_data_row(ws, row, ["Income before tax", "='Income Statement'!B28", "='Income Statement'!C28", "='Income Statement'!D28"])
    row += 1  # row 9
    apply_data_row(ws, row, ["Adjustments for non-cash items", 13856, 14955, 14909])
    row += 1  # row 10
    apply_data_row(ws, row, ["Cash flow before working capital changes", "=B8+B9", "=C8+C9", "=D8+D9"], 'subtotal')
    row += 1  # row 11
    apply_data_row(ws, row, ["Change in inventory", -3439, -3760, 2187])
    row += 1  # row 12
    apply_data_row(ws, row, ["Change in receivables", -10532, -7111, 11601])
    row += 1  # row 13
    apply_data_row(ws, row, ["Change in payables", 10899, 1721, -7695])
    row += 1  # row 14
    apply_data_row(ws, row, ["Net cash from operating activities", "=B10+SUM(B11:B13)", "=C10+SUM(C11:C13)", "=D10+SUM(D11:D13)"], 'subtotal')
    row += 2  # row 16

    # Investing activities
    apply_data_row(ws, row, ["Cash flows from investing activities", "", "", ""], 'section')
    row += 1  # row 17
    apply_data_row(ws, row, ["Investment in intangible assets", -4414, -6360, -3775])
    row += 1  # row 18
    apply_data_row(ws, row, ["Investment in tangible assets", -71, 2809, -5194])
    row += 1  # row 19
    apply_data_row(ws, row, ["Change in financial assets", -5000, 0, 0])
    row += 1  # row 20
    apply_data_row(ws, row, ["Net cash from investing activities", "=SUM(B17:B19)", "=SUM(C17:C19)", "=SUM(D17:D19)"], 'subtotal')
    row += 2  # row 22

    # Financing activities
    apply_data_row(ws, row, ["Cash flows from financing activities", "", "", ""], 'section')
    row += 1  # row 23
    apply_data_row(ws, row, ["Proceeds from share issue", 78089, -378, 0])
    row += 1  # row 24
    apply_data_row(ws, row, ["Share issue costs", -4977, 260, 0])
    row += 1  # row 25
    apply_data_row(ws, row, ["Net cash from financing activities", "=SUM(B23:B24)", "=SUM(C23:C24)", "=SUM(D23:D24)"], 'subtotal')
    row += 2  # row 27

    # Net change in cash
    apply_data_row(ws, row, ["Net change in cash", "=B14+B20+B25", "=C14+C20+C25", "=D14+D20+D25"], 'total')
    row += 1  # row 28
    apply_data_row(ws, row, ["Cash at beginning of period", 16467, 71822, 53984])
    row += 1  # row 29
    apply_data_row(ws, row, ["FX effect on cash", -158, -24, -27])
    row += 1  # row 30
    apply_data_row(ws, row, ["Cash at end of period", "=SUM(B27:B29)", "=SUM(C27:C29)", "=SUM(D27:D29)"], 'total')
    row += 3

    # Source
    ws.cell(row=row, column=1, value="Source: Freemelt Interim Reports Q1-Q3 2025, page 20").font = SOURCE_FONT

    # Kolumnbredder
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    ws.freeze_panes = 'A6'

    # Ta bort stödlinjer (gridlines)
    ws.sheet_view.showGridLines = False

    return wb


def create_workbook_from_tables(config: dict):
    """
    Skapa Excel-arbetsbok från extraherade RAW tabeller
    Auto-detekterar formatering (headers, sections, subtotals, totals)
    """
    
    wb = Workbook()
    wb.remove(wb.active)
    
    company_name = config['company_name']
    year = config['year']
    
    # ============================================
    # FLIK 1: RESULTATRÄKNING
    # ============================================
    ws = wb.create_sheet("Income Statement")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "Consolidated Income Statement"
    ws['A2'].font = Font(name='Arial', size=10, color=GS_DARK_GRAY)
    
    income_statement = config.get('income_statement', [])
    _populate_sheet(ws, income_statement, company_name, year, "Income Statement")
    
    # ============================================
    # FLIK 2: BALANSRÄKNING
    # ============================================
    ws = wb.create_sheet("Balance Sheet")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "Consolidated Balance Sheet"
    ws['A2'].font = Font(name='Arial', size=10, color=GS_DARK_GRAY)
    
    balance_sheet = config.get('balance_sheet', [])
    _populate_sheet(ws, balance_sheet, company_name, year, "Balance Sheet")
    
    # ============================================
    # FLIK 3: KASSAFLÖDESANALYS
    # ============================================
    ws = wb.create_sheet("Cash Flow")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "Consolidated Statement of Cash Flows"
    ws['A2'].font = Font(name='Arial', size=10, color=GS_DARK_GRAY)
    
    cash_flow = config.get('cash_flow', [])
    _populate_sheet(ws, cash_flow, company_name, year, "Cash Flow")
    
    return wb


def _populate_sheet(ws, table_data: list, company_name: str, year: int, sheet_type: str):
    """
    Fylla ett blad med extraherad tabelldata
    Auto-detekterar row-typ (header, section, subtotal, total, data)
    """
    
    if not table_data:
        return
    
    # Räkna värdekolumner
    num_value_cols = 3
    if len(table_data) > 0:
        num_value_cols = len(table_data[0]) - 1
    
    # Sätt kolumnbredder
    ws.column_dimensions['A'].width = 36
    for col in range(2, 2 + num_value_cols):
        col_letter = chr(64 + col)
        ws.column_dimensions[col_letter].width = 14
    
    current_row = 4
    
    # Header row
    if len(table_data) > 0:
        first_row = table_data[0]
        is_header = _is_header_row(first_row)
        
        if is_header:
            for col, value in enumerate(first_row, 1):
                cell = ws.cell(row=4, column=col, value=value)
                if col == 1:
                    cell.value = "SEK '000"
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN
                cell.border = HEADER_BORDER
            
            current_row = 5
            start_data = 1
        else:
            default_headers = ["SEK '000"] + ["Q1", "Q2", "Q3"][:num_value_cols]
            apply_header_row(ws, 4, default_headers)
            current_row = 5
            start_data = 0
    
    # Subheader
    ws['A5'] = ""
    subheaders = ["Jan-Mar", "Apr-Jun", "Jul-Sep"]
    if "Balance" in sheet_type:
        subheaders = ["31-Mar", "30-Jun", "30-Sep"]
    
    for col, sub in enumerate(subheaders[:num_value_cols], 2):
        ws.cell(row=5, column=col, value=sub).font = SUBHEADER_FONT
        ws.cell(row=5, column=col).alignment = RIGHT_ALIGN
    
    current_row = 7
    
    # Data rows
    for table_row in table_data:
        if not table_row or (len(table_row) == 1 and not table_row[0]):
            continue
        
        label = str(table_row[0]) if table_row else ""
        values = table_row[1:] if len(table_row) > 1 else []
        
        row_type = _detect_row_type(label, values)
        
        row_values = [label] + values
        apply_data_row(ws, current_row, row_values, row_type)
        
        current_row += 1
    
    # Source row
    current_row += 2
    ws.cell(row=current_row, column=1, value=f"Source: {company_name} Q1-Q3 {year}").font = SOURCE_FONT
    
    ws.freeze_panes = 'A6'
    ws.sheet_view.showGridLines = False


def _is_header_row(row: list) -> bool:
    """Detektera om rad är header"""
    if not row or len(row) < 2:
        return False
    
    label = str(row[0]).lower()
    header_keywords = ["sek", "item", "description", "account", "q1", "q2", "q3", "jan", "mar"]
    return any(keyword in label for keyword in header_keywords)


def _detect_row_type(label: str, values: list) -> str:
    """Auto-detektera rad-typ"""
    
    label_lower = label.lower().strip()
    
    # Total eller subtotal
    if any(word in label_lower for word in ["total", "sum", "subtotal"]):
        if any(word in label_lower for word in ["total assets", "total equity", "total liabilities", "net income", "net change"]):
            return 'total'
        else:
            return 'subtotal'
    
    # Section header
    if all(v == "" or v is None for v in values):
        if label and label.strip():
            return 'section'
    
    return 'data'


if __name__ == "__main__":
    wb = create_workbook()
    output_path = "/Users/oskarhornell/Github/databok-claude/Freemelt_Kvartalsrapporter_2025.xlsx"
    wb.save(output_path)
    print(f"Excel-fil skapad: {output_path}")
    print("\nInvestment Bank formatering (Goldman Sachs-stil):")
    print("- Arial typsnitt, 9pt")
    print("- Mörkblå headers (#1F3864)")
    print("- Negativa tal i parenteser")
    print("- Dubbla linjer under totaler")
    print("- Cross-sheet länkar till Income Statement")
    print("\nFlikar:")
    print("1. Income Statement")
    print("2. Balance Sheet")
    print("3. Cash Flow")
