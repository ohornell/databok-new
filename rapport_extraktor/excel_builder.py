"""
Excel-databok byggare med Investment Bank-formatering.
Skapar professionellt formaterade finansiella rapporter.

Förenklad version - endast resultaträkning, balansräkning och kassaflöde.
Inkluderar AI-driven radnormalisering för att matcha liknande radnamn mellan kvartal.
"""

import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def sanitize_sheet_name(name: str) -> str:
    """Sanera fliknamn för Excel (tar bort ogiltiga tecken)."""
    if not name:
        return "Sheet"
    # Excel tillåter inte: / \ * ? : [ ]
    invalid_chars = r'[/\\*?:\[\]]'
    sanitized = re.sub(invalid_chars, ' ', name)
    # Ta bort dubbla mellanslag
    sanitized = re.sub(r'\s+', ' ', sanitized).strip()
    # Max 31 tecken
    return sanitized[:31]


def normalize_row_name(name: str) -> str:
    """Normalisera radnamn för matchning mellan perioder."""
    if not name:
        return ""
    # Lowercase, ta bort extra whitespace
    normalized = name.lower().strip()
    # Ta bort vanliga variationer
    normalized = normalized.replace("  ", " ")
    # Ta bort parenteser med innehåll (t.ex. "(MSEK)")
    normalized = re.sub(r'\s*\([^)]*\)', '', normalized)
    return normalized


def map_table_type(table: dict) -> str:
    """
    Mappa tabelltyp, inklusive quarterly → rätt typ baserat på titel.

    Quarterly-tabeller (som "Koncernens resultaträkningar per kvartal")
    mappas till sin underliggande typ (income_statement, kpi, etc.)
    """
    table_type = table.get("type", "other")
    if table_type == "quarterly":
        title = table.get("title", "").lower()
        if "resultat" in title or "income" in title:
            return "income_statement"
        elif "balans" in title or "ställning" in title or "balance" in title:
            return "balance_sheet"
        elif "kassaflöde" in title or "cash" in title:
            return "cash_flow"
        elif "nyckeltal" in title or "kpi" in title:
            return "kpi"
        elif "segment" in title:
            return "segment"
        else:
            return "other"
    return table_type


def extract_year_from_column(col_name: str) -> int:
    """
    Extrahera årtal från kolumnnamn för sortering.

    Exempel:
        "Juli-september 2025" -> 2025
        "Januari-mars 2024" -> 2024
        "Helår 2024" -> 2024
        "Q3 2025" -> 2025

    Returns:
        Årtal som int, eller 0 om inget hittades
    """
    if not col_name:
        return 0
    # Sök efter 4-siffrigt årtal (2020-2030)
    match = re.search(r'20[2-3]\d', str(col_name))
    return int(match.group()) if match else 0


def determine_current_year(data_list: list[dict]) -> int:
    """
    Bestäm dynamiskt vilket år som är "current" baserat på datan.

    Returnerar det SENASTE året som förekommer i perioderna.
    """
    years = set()
    for item in data_list:
        period = item.get("metadata", {}).get("period", "")
        year = extract_year_from_column(period)
        if year:
            years.add(year)

        # Kolla även kolumnrubriker
        for table in item.get("tables", []):
            for col in table.get("columns", []):
                year = extract_year_from_column(col)
                if year:
                    years.add(year)

    return max(years) if years else 2025  # Fallback till 2025 om inget hittas


# ============================================
# INVESTMENT BANK STYLE GUIDE
# ============================================

# Färgpalett (Goldman Sachs-inspirerad)
GS_NAVY = "1F3864"
GS_LIGHT_BLUE = "D6DCE4"
GS_LIGHT_GRAY = "F2F2F2"
GS_DARK_GRAY = "404040"
GS_BLACK = "000000"

# Färgkodning för data
COLOR_HARDCODED = "0000FF"  # Blå - hårdkodade värden

# Fonter
TITLE_FONT = Font(name='Arial', size=11, bold=True, color=GS_NAVY)
TABLE_TITLE_FONT = Font(name='Arial', size=11, bold=True, color=GS_NAVY)  # Tabellrubriker
SUBTITLE_FONT = Font(name='Arial', size=10, color=GS_DARK_GRAY)
HEADER_FONT = Font(name='Arial', size=9, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name='Arial', size=8, italic=True, color=GS_DARK_GRAY)
SECTION_FONT = Font(name='Arial', size=9, bold=True, color=GS_NAVY)
LABEL_FONT = Font(name='Arial', size=9, color=GS_DARK_GRAY)
DATA_FONT = Font(name='Arial', size=9, color=COLOR_HARDCODED)
TOTAL_FONT = Font(name='Arial', size=9, bold=True, color=GS_BLACK)
SUBTOTAL_FONT = Font(name='Arial', size=9, bold=True, color=GS_DARK_GRAY)
SOURCE_FONT = Font(name='Arial', size=7, italic=True, color="808080")

# Fyllningar
HEADER_FILL = PatternFill(start_color=GS_NAVY, end_color=GS_NAVY, fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color=GS_LIGHT_BLUE, end_color=GS_LIGHT_BLUE, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=GS_LIGHT_GRAY, end_color=GS_LIGHT_GRAY, fill_type="solid")

# Ramar
thin_side = Side(style='thin', color=GS_DARK_GRAY)
medium_side = Side(style='medium', color=GS_DARK_GRAY)
double_side = Side(style='double', color=GS_BLACK)

HEADER_BORDER = Border(bottom=medium_side)
SECTION_BORDER = Border(bottom=thin_side)
SUBTOTAL_BORDER = Border(top=thin_side, bottom=thin_side)
TOTAL_BORDER = Border(top=thin_side, bottom=double_side)
NO_BORDER = Border()

# Alignment
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', indent=1)
INDENT_ALIGN = Alignment(horizontal='left', vertical='center', indent=2)

# Nummerformat
NUMBER_FORMAT = '#,##0_);(#,##0);"-"_)'
PERCENT_FORMAT = '0.0%_);(0.0%)'

# Font för periodavdelare
PERIOD_SEPARATOR_FONT = Font(name='Arial', size=12, bold=True, color="FFFFFF")
PERIOD_SEPARATOR_FILL = PatternFill(start_color=GS_NAVY, end_color=GS_NAVY, fill_type="solid")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def write_period_separator(ws, row: int, period: str, num_cols: int = 5, is_multi_period: bool = True) -> int:
    """
    Skriv en tydlig periodavdelare i Excel.
    Endast om det är multi-period export.

    Args:
        ws: Worksheet
        row: Rad att börja på
        period: Periodnamn (t.ex. "Q1 2025")
        num_cols: Antal kolumner att slå ihop
        is_multi_period: Om True, skriv avdelare. Om False, returnera direkt.

    Returns:
        Nästa lediga rad
    """
    # Hoppa över avdelare för enskilt kvartal
    if not is_multi_period:
        return row

    # Övre linje
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PERIOD_SEPARATOR_FILL
    row += 1

    # Period-text (centrerad, stor font)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=period)
    cell.font = PERIOD_SEPARATOR_FONT
    cell.fill = PERIOD_SEPARATOR_FILL
    cell.alignment = CENTER_ALIGN
    row += 1

    # Undre linje
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PERIOD_SEPARATOR_FILL
    row += 1

    # Tom rad efter
    row += 1

    return row


def sort_by_period(data: list[dict]) -> list[dict]:
    """
    Sortera extraherad data kronologiskt efter period.
    Hanterar format som Q1 2025, Q2 2024, etc.
    """
    def period_key(item):
        period = item.get("metadata", {}).get("period", "")
        # Extrahera Q-nummer och år
        match = re.search(r'Q(\d)\s*(\d{4})', period)
        if match:
            quarter = int(match.group(1))
            year = int(match.group(2))
            return (year, quarter)
        return (0, 0)

    return sorted(data, key=period_key)


def collect_all_rows(data_list: list[dict], data_key: str) -> list[str]:
    """
    Samla alla unika radnamn från alla perioder med smart ordning.

    Algoritm:
    1. Använd första kvartalets ordning som bas
    2. När nya rader dyker upp i senare kvartal, försök placera dem
       på rätt position baserat på omgivande rader
    3. Normalisera radnamn för jämförelse (t.ex. "receivables" -> "receivable")
    """
    if not data_list:
        return []

    # Samla alla rader med normaliserade namn för jämförelse
    # Key: normaliserat namn, Value: (originalnamn, första_index_per_period)
    seen_normalized = {}

    # Bygg ordnad lista baserad på alla perioders ordning
    ordered_rows = []

    for period_idx, item in enumerate(data_list):
        rows = item.get(data_key, [])
        prev_normalized = None

        for row_idx, row in enumerate(rows):
            row_name = row.get("rad") or row.get("namn") or row.get("region", "")
            if not row_name:
                continue

            norm = normalize_row_name(row_name)

            if norm not in seen_normalized:
                # Ny rad - behöver placeras
                seen_normalized[norm] = row_name

                if period_idx == 0:
                    # Första perioden - lägg till direkt
                    ordered_rows.append(row_name)
                else:
                    # Senare period - försök placera efter föregående rad
                    if prev_normalized and prev_normalized in seen_normalized:
                        # Hitta positionen för föregående rad
                        prev_orig = seen_normalized[prev_normalized]
                        try:
                            prev_pos = ordered_rows.index(prev_orig)
                            ordered_rows.insert(prev_pos + 1, row_name)
                        except ValueError:
                            # Föregående rad hittades inte, lägg till sist
                            ordered_rows.append(row_name)
                    else:
                        # Ingen föregående rad att referera till, lägg till sist
                        ordered_rows.append(row_name)

            prev_normalized = norm

    return ordered_rows


def detect_row_type(row_data: dict, row_name: str) -> str:
    """
    Detektera radtyp baserat på data och namn.
    """
    # Explicit typ från extraktionen
    if row_data.get("typ") == "total":
        return "total"
    if row_data.get("typ") == "subtotal":
        return "subtotal"

    # Detektera baserat på nyckelord
    name_lower = row_name.lower()

    total_keywords = ["summa", "total", "netto", "resultat efter"]
    if any(kw in name_lower for kw in total_keywords):
        if "summa" in name_lower or "total" in name_lower:
            return "total" if "tillgångar" in name_lower or "skulder" in name_lower else "subtotal"

    return "data"


def apply_row_style(ws, row_num: int, num_cols: int, row_type: str, row_name: str):
    """
    Applicera stil på en rad baserat på typ.
    """
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)

        if row_type == "section":
            cell.font = SECTION_FONT
            cell.border = SECTION_BORDER
            cell.alignment = LEFT_ALIGN
        elif row_type == "subtotal":
            cell.fill = SUBTOTAL_FILL
            cell.border = SUBTOTAL_BORDER
            cell.font = SUBTOTAL_FONT if col == 1 else Font(name='Arial', size=9, bold=True, color=COLOR_HARDCODED)
            cell.alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN
            if col > 1:
                cell.number_format = NUMBER_FORMAT
        elif row_type == "total":
            cell.fill = TOTAL_FILL
            cell.border = TOTAL_BORDER
            cell.font = TOTAL_FONT if col == 1 else Font(name='Arial', size=9, bold=True, color=COLOR_HARDCODED)
            cell.alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN
            if col > 1:
                cell.number_format = NUMBER_FORMAT
        else:  # data
            cell.border = NO_BORDER
            if col == 1:
                cell.font = LABEL_FONT
                cell.alignment = INDENT_ALIGN
            else:
                cell.font = DATA_FONT
                cell.alignment = RIGHT_ALIGN
                cell.number_format = NUMBER_FORMAT


def populate_financial_sheet(
    ws,
    data_list: list[dict],
    data_key: str,
    periods: list[str],
    company_name: str
):
    """
    Fyll ett finansiellt blad med data från alla perioder.
    """
    num_periods = len(periods)

    # Titel
    ws.merge_cells(f'A1:{get_column_letter(num_periods + 1)}1')
    ws['A1'] = company_name.upper()
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN

    # Undertitel baserad på data_key (förenklad version)
    titles = {
        "resultatrakning": "Resultaträkning",
        "balansrakning": "Balansräkning",
        "kassaflodesanalys": "Kassaflödesanalys",
    }
    ws.merge_cells(f'A2:{get_column_letter(num_periods + 1)}2')
    ws['A2'] = titles.get(data_key, data_key.replace("_", " ").title())
    ws['A2'].font = SUBTITLE_FONT

    # Header-rad
    headers = [""]
    for item in data_list:
        period = item.get("metadata", {}).get("period", "?")
        headers.append(period)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN if col == 1 else RIGHT_ALIGN
        cell.border = HEADER_BORDER

    # Valuta i första cellen
    valuta = data_list[0].get("metadata", {}).get("valuta", "TSEK") if data_list else "TSEK"
    ws.cell(row=4, column=1, value=valuta)

    # Samla alla radnamn
    all_rows = collect_all_rows(data_list, data_key)

    # Skriv data
    current_row = 6
    row_name_normalized = normalize_row_name  # Referens för snabbare anrop

    for row_name in all_rows:
        # Hämta värden för varje period
        values = [row_name]
        row_data = {}
        target_norm = row_name_normalized(row_name)

        for item in data_list:
            rows = item.get(data_key, [])
            value = None
            for r in rows:
                r_name = r.get("rad") or r.get("namn") or r.get("region", "")
                # Använd normaliserad jämförelse för att matcha liknande radnamn
                if row_name_normalized(r_name) == target_norm:
                    value = r.get("varde")
                    row_data = r
                    break
            values.append(value)

        # Skriv rad
        for col, val in enumerate(values, 1):
            ws.cell(row=current_row, column=col, value=val)

        # Detektera och applicera stil
        row_type = detect_row_type(row_data, row_name)
        apply_row_style(ws, current_row, num_periods + 1, row_type, row_name)

        current_row += 1

    # Källa
    current_row += 2
    ws.cell(row=current_row, column=1, value=f"Källa: {company_name} kvartalsrapporter").font = SOURCE_FONT

    # Kolumnbredder
    ws.column_dimensions['A'].width = 36
    for col in range(2, num_periods + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Frys rubriker
    ws.freeze_panes = 'A5'

    # Dölj gridlines
    ws.sheet_view.showGridLines = False


def populate_notes_sheet(ws, data_list: list[dict], company_name: str):
    """
    Speciell hantering för noter som har annan struktur.
    """
    ws.merge_cells('A1:D1')
    ws['A1'] = company_name.upper()
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:D2')
    ws['A2'] = "Noter"
    ws['A2'].font = SUBTITLE_FONT

    current_row = 4

    # Samla alla noter från alla perioder
    all_notes = {}
    for item in data_list:
        period = item.get("metadata", {}).get("period", "?")
        for note in item.get("noter", []):
            note_num = note.get("nummer", 0)
            if note_num not in all_notes:
                all_notes[note_num] = {
                    "titel": note.get("titel", ""),
                    "perioder": {}
                }
            all_notes[note_num]["perioder"][period] = note

    # Skriv noter
    for note_num in sorted(all_notes.keys()):
        note_info = all_notes[note_num]

        # Not-rubrik
        ws.cell(row=current_row, column=1, value=f"Not {note_num}: {note_info['titel']}")
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        current_row += 1

        # Tabeller från noten (ta från senaste period)
        if note_info["perioder"]:
            latest_note = list(note_info["perioder"].values())[-1]
            for table in latest_note.get("tabeller", []):
                # Tabellrubrik
                ws.cell(row=current_row, column=1, value=table.get("rubrik", ""))
                ws.cell(row=current_row, column=1).font = SUBTOTAL_FONT
                current_row += 1

                # Tabellrader
                for rad in table.get("rader", []):
                    ws.cell(row=current_row, column=1, value=rad.get("rad", ""))
                    ws.cell(row=current_row, column=2, value=rad.get("varde"))
                    ws.cell(row=current_row, column=1).font = LABEL_FONT
                    ws.cell(row=current_row, column=2).font = DATA_FONT
                    ws.cell(row=current_row, column=2).number_format = NUMBER_FORMAT
                    current_row += 1

        current_row += 1

    # Kolumnbredder
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 14

    ws.sheet_view.showGridLines = False


def populate_dynamic_table_sheet(
    ws,
    data_list: list[dict],
    table_type: str,
    company_name: str
):
    """
    Fyll ett blad med dynamiska tabeller från full extraktion.

    Hanterar tables-formatet: {"title", "page", "type", "columns", "rows"}
    Varje tabell visas separat med alla sina kolumner.
    """
    # Hitta alla tabeller av denna typ från alla perioder
    # Använd map_table_type för att inkludera quarterly-tabeller
    all_tables = []
    for item in data_list:
        period = item.get("metadata", {}).get("period", "?")
        for table in item.get("tables", []):
            if map_table_type(table) == table_type:
                all_tables.append({
                    "period": period,
                    "table": table
                })

    if not all_tables:
        return

    # Titel
    type_titles = {
        "income_statement": "Resultaträkning",
        "balance_sheet": "Balansräkning",
        "cash_flow": "Kassaflödesanalys",
        "kpi": "Nyckeltal",
        "segment": "Segmentdata",
        "other": "Övriga tabeller",
    }

    current_row = 1

    # Bolagsnamn som huvudrubrik
    ws['A1'] = company_name.upper()
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN
    current_row = 3

    # Kolla om det är multi-period (för periodavdelare)
    is_multi_period = len(data_list) > 1
    current_period = None

    # Skriv ut varje tabell separat
    for table_idx, table_info in enumerate(all_tables):
        table = table_info["table"]
        period = table_info["period"]

        # Lägg till periodavdelare om ny period (endast multi-period)
        if period != current_period:
            current_row = write_period_separator(ws, current_row, period, num_cols=8, is_multi_period=is_multi_period)
            current_period = period

        # Tabellens titel med sidnummer
        title = table.get("title", type_titles.get(table_type, "Tabell"))
        page = table.get("page")
        if page:
            title_with_page = f"{title} (s. {page})"
        else:
            title_with_page = title
        ws.cell(row=current_row, column=1, value=title_with_page)
        ws.cell(row=current_row, column=1).font = TABLE_TITLE_FONT
        current_row += 1

        # Kolumnrubriker från tabellen
        columns = table.get("columns", [])
        # Första kolumnen är tom (för radnamn), resten är värdekolumner
        # Hoppa över första kolumnen om den är tom/bara beskrivning
        value_columns = columns[1:] if columns and columns[0] in ["", "MSEK", "TSEK", "SEK"] else columns

        # Filtrera bort "Not"/"Note"-kolumner (inte numeriska värden)
        # not_col_indices är relativt till value_columns (som börjar efter label-kolumnen)
        not_col_indices = set(i for i, c in enumerate(value_columns) if str(c).lower() in ["not", "note", "notes"])
        if not_col_indices:
            value_columns = [c for i, c in enumerate(value_columns) if i not in not_col_indices]

        # Header-rad
        ws.cell(row=current_row, column=1, value="").font = HEADER_FONT
        ws.cell(row=current_row, column=1).fill = HEADER_FILL
        ws.cell(row=current_row, column=1).border = HEADER_BORDER

        for col_idx, col_name in enumerate(value_columns, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = RIGHT_ALIGN
            cell.border = HEADER_BORDER

        current_row += 1

        # Data-rader
        rows = table.get("rows", [])
        num_cols = len(value_columns) + 1

        for row_data in rows:
            label = row_data.get("label", "")
            values = row_data.get("values", [])
            row_type = row_data.get("type", "data")

            # Radnamn
            ws.cell(row=current_row, column=1, value=label)

            # Värden - hoppa över Not-kolumnens index
            # values[0] är alltid null (label), values[1:] motsvarar value_columns innan filtrering
            # Vi måste filtrera values[1:] med samma not_col_indices
            filtered_values = [v for i, v in enumerate(values[1:]) if i not in not_col_indices] if not_col_indices else values[1:]
            for val_idx, value in enumerate(filtered_values):
                if val_idx + 2 <= num_cols:
                    ws.cell(row=current_row, column=val_idx + 2, value=value)

            # Applicera stil
            apply_row_style(ws, current_row, num_cols, row_type, label)
            current_row += 1

        # Mellanrum mellan tabeller
        current_row += 2

    # Källa
    ws.cell(row=current_row, column=1, value=f"Källa: {company_name} kvartalsrapporter").font = SOURCE_FONT

    # Kolumnbredder
    ws.column_dimensions['A'].width = 45
    for col in range(2, 10):  # Max 8 värdekolumner
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.sheet_view.showGridLines = False


def populate_sections_sheet(ws, data_list: list[dict], section_title: str, company_name: str):
    """
    Fyll ett blad med textsektioner från full extraktion.
    Visar samma sektion från alla kvartal.
    """
    # Fonter
    TEXT_FONT = Font(name='Arial', size=10, color=GS_DARK_GRAY)
    BULLET_FONT = Font(name='Arial', size=10, color=GS_DARK_GRAY)
    SUBHEADER_FONT = Font(name='Arial', size=10, bold=True, color=GS_DARK_GRAY)

    def is_subheader(line: str) -> bool:
        """Kolla om en rad är en underrubrik (kort, utan bullet, ej siffra)."""
        if len(line) > 50:  # För lång för att vara rubrik
            return False
        if line.startswith(('• ', '- ', '* ', '– ')):  # Bullet
            return False
        if len(line) > 2 and line[0].isdigit() and line[1] in '.):':  # Numrerad
            return False
        # Kort text som ser ut som rubrik
        words = line.split()
        if len(words) <= 6 and not line.endswith(('.', ',')):
            return True
        return False

    def content_similarity(c1: str, c2: str) -> float:
        """Beräkna likhet mellan två texter (0-1)."""
        # Ta första 200 tecken för snabb jämförelse
        s1 = c1[:200].lower().replace(' ', '')
        s2 = c2[:200].lower().replace(' ', '')
        if not s1 or not s2:
            return 0
        # Enkel Jaccard-liknande metric
        common = sum(1 for c in s1 if c in s2)
        return common / max(len(s1), len(s2))

    # Hitta alla sektioner med denna titel
    all_sections = []
    for item in data_list:
        period = item.get("metadata", {}).get("period", "?")
        for section in item.get("sections", []):
            if section.get("title") == section_title:
                all_sections.append({
                    "period": period,
                    "section": section
                })

    if not all_sections:
        return

    # Deduplicera sektioner med liknande innehåll (>80% likhet)
    unique_sections = []
    for section_info in all_sections:
        content = section_info["section"].get("content", "")
        is_duplicate = False
        for existing in unique_sections:
            existing_content = existing["section"].get("content", "")
            if content_similarity(content, existing_content) > 0.8:
                is_duplicate = True
                break
        if not is_duplicate:
            unique_sections.append(section_info)

    # Bolagsnamn som huvudrubrik (samma som tabeller)
    ws['A1'] = company_name.upper()
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN

    current_row = 3

    # Kolla om det är multi-period (för periodavdelare)
    is_multi_period = len(data_list) > 1
    current_period = None

    for section_info in unique_sections:
        period = section_info["period"]
        section = section_info["section"]

        # Lägg till periodavdelare om ny period (endast multi-period)
        if period != current_period:
            current_row = write_period_separator(ws, current_row, period, num_cols=1, is_multi_period=is_multi_period)
            current_period = period

        # Sektionens titel med sidnummer (samma format som tabeller)
        page = section.get("page")
        if page:
            title_with_page = f"{section_title} (s. {page})"
        else:
            title_with_page = section_title
        ws.cell(row=current_row, column=1, value=title_with_page)
        ws.cell(row=current_row, column=1).font = TABLE_TITLE_FONT
        current_row += 1

        # Textinnehåll - behåll styckeindelning och punktlistor
        content = section.get("content", "")

        # Dela upp på stycken (dubbla radbrytningar)
        paragraphs = content.split('\n\n')

        for para in paragraphs:
            para = para.strip()
            if not para:
                continue

            lines = para.split('\n')
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Identifiera radtyp
                is_bullet = False
                is_sub = False

                if line.startswith(('• ', '- ', '* ', '– ')):
                    is_bullet = True
                elif len(line) > 2 and line[0].isdigit() and line[1] in '.):':
                    is_bullet = True
                elif is_subheader(line):
                    is_sub = True

                # Välj font
                if is_sub:
                    font = SUBHEADER_FONT
                elif is_bullet:
                    font = BULLET_FONT
                else:
                    font = TEXT_FONT

                # Radbryt långa rader (max 120 tecken)
                if len(line) > 120:
                    words = line.split()
                    wrapped_lines = []
                    current_line = ""
                    indent = "   " if is_bullet else ""

                    for word in words:
                        test_line = current_line + " " + word if current_line else word
                        if len(test_line) <= 120:
                            current_line = test_line
                        else:
                            if current_line:
                                wrapped_lines.append(current_line)
                            current_line = indent + word if wrapped_lines else word
                    if current_line:
                        wrapped_lines.append(current_line)

                    for wline in wrapped_lines:
                        cell = ws.cell(row=current_row, column=1, value=wline)
                        cell.font = font
                        cell.alignment = Alignment(wrap_text=False, vertical='top')
                        current_row += 1
                else:
                    cell = ws.cell(row=current_row, column=1, value=line)
                    cell.font = font
                    cell.alignment = Alignment(wrap_text=False, vertical='top')
                    current_row += 1

            # Tom rad mellan stycken
            current_row += 1

        current_row += 1  # Extra mellanrum efter sektion

    # Källa (samma som tabeller)
    ws.cell(row=current_row, column=1, value=f"Källa: {company_name} kvartalsrapporter").font = SOURCE_FONT

    # Kolumnbredd
    ws.column_dimensions['A'].width = 120
    ws.sheet_view.showGridLines = False


def populate_charts_sheet(ws, data_list: list[dict], company_name: str):
    """
    Fyll ett blad med grafdata från full extraktion.
    Varje graf visas som en tabell med datapunkter OCH som en riktig Excel-graf.
    Goldman Sachs Investment Banking-stil med professionell formatering.
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

    # Goldman Sachs färgpalett för grafer (hex utan #)
    GS_NAVY = "1F3864"
    GS_BLUE = "4472C4"
    GS_LIGHT_BLUE = "8FAADC"
    GS_GRAY = "7F7F7F"

    # Samla alla grafer
    all_charts = []
    for item in data_list:
        period = item.get("metadata", {}).get("period", "?")
        for chart in item.get("charts", []):
            all_charts.append({
                "period": period,
                "chart": chart
            })

    if not all_charts:
        return

    # Titel
    ws['A1'] = company_name.upper()
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = LEFT_ALIGN

    ws['A2'] = "Extraherade grafer och diagram"
    ws['A2'].font = SUBTITLE_FONT

    current_row = 4
    chart_count = 0

    # Kolla om det är multi-period (för periodavdelare)
    is_multi_period = len(data_list) > 1
    current_period = None

    for chart_info in all_charts:
        period = chart_info["period"]
        chart = chart_info["chart"]

        # Lägg till periodavdelare om ny period (endast multi-period)
        if period != current_period:
            current_row = write_period_separator(ws, current_row, period, num_cols=3, is_multi_period=is_multi_period)
            current_period = period

        # Graf-rubrik med IB-stil
        title = chart.get("title", "Graf")
        estimated = chart.get("estimated", True)

        # Rubrikrad
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        ws.cell(row=current_row, column=1).border = SECTION_BORDER
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        current_row += 1

        # Metadata-rad
        meta_parts = [period]
        if chart.get("y_axis"):
            meta_parts.append(chart["y_axis"])
        if estimated:
            meta_parts.append("Uppskattade värden")
        else:
            meta_parts.append("Exakta värden")

        ws.cell(row=current_row, column=1, value=" | ".join(meta_parts))
        ws.cell(row=current_row, column=1).font = SOURCE_FONT
        current_row += 1

        # Datapunkter som tabell
        data_points = chart.get("data_points", [])
        data_start_row = current_row
        if data_points:
            # Header med IB-stil
            ws.cell(row=current_row, column=1, value="")
            ws.cell(row=current_row, column=2, value="Värde")
            for col in [1, 2]:
                cell = ws.cell(row=current_row, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = RIGHT_ALIGN if col == 2 else LEFT_ALIGN
                cell.border = HEADER_BORDER
            current_row += 1

            # Data med IB-stil
            for dp in data_points:
                ws.cell(row=current_row, column=1, value=dp.get("label", ""))
                ws.cell(row=current_row, column=1).font = LABEL_FONT
                ws.cell(row=current_row, column=1).alignment = LEFT_ALIGN

                val_cell = ws.cell(row=current_row, column=2, value=dp.get("value"))
                val_cell.font = DATA_FONT
                val_cell.alignment = RIGHT_ALIGN
                # Använd alltid nummerformat (inte procent)
                val_cell.number_format = NUMBER_FORMAT
                current_row += 1

            data_end_row = current_row - 1

            # Skapa Excel-graf baserat på chart_type
            chart_type = chart.get("chart_type", "bar")

            if chart_type == "pie":
                # Cirkeldiagram - Goldman Sachs stil
                from openpyxl.chart.series import DataPoint
                excel_chart = PieChart()
                labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
                data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                excel_chart.add_data(data_ref, titles_from_data=True)
                excel_chart.set_categories(labels)

                # Goldman Sachs färgpalett för pie-sektorer
                pie_colors = [GS_NAVY, GS_BLUE, GS_LIGHT_BLUE, GS_GRAY, "A5A5A5", "D9D9D9"]

                # Datapunktetiketter
                excel_chart.dataLabels = DataLabelList()
                excel_chart.dataLabels.showPercent = True
                excel_chart.dataLabels.showVal = False
                excel_chart.dataLabels.showCatName = True
                excel_chart.dataLabels.showSerName = False

                # Sätt färger på varje sektor
                if excel_chart.series:
                    series = excel_chart.series[0]
                    num_points = len(data_points)
                    for i in range(num_points):
                        pt = DataPoint(idx=i)
                        color = pie_colors[i % len(pie_colors)]
                        pt.graphicalProperties.solidFill = color
                        pt.graphicalProperties.line.noFill = True
                        series.data_points.append(pt)

            elif chart_type == "line":
                # Linjediagram - Goldman Sachs stil
                excel_chart = LineChart()
                excel_chart.style = 10  # Enkel stil
                labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
                data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                excel_chart.add_data(data_ref, titles_from_data=True)
                excel_chart.set_categories(labels)

                # Y-axel - ta bort gridlines
                excel_chart.y_axis.majorGridlines = None
                excel_chart.y_axis.delete = False

                # Datapunktetiketter
                excel_chart.dataLabels = DataLabelList()
                excel_chart.dataLabels.showVal = True

            elif chart_type == "area":
                # Ytdiagram (area chart) - Goldman Sachs stil
                from openpyxl.chart import AreaChart
                excel_chart = AreaChart()
                excel_chart.style = 10  # Enkel stil
                excel_chart.grouping = "standard"
                labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
                data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                excel_chart.add_data(data_ref, titles_from_data=True)
                excel_chart.set_categories(labels)

                # Y-axel - ta bort gridlines
                excel_chart.y_axis.majorGridlines = None
                excel_chart.y_axis.delete = False

                # Datapunktetiketter
                excel_chart.dataLabels = DataLabelList()
                excel_chart.dataLabels.showVal = True

            else:
                # Stapeldiagram (bar/default) - Goldman Sachs stil
                excel_chart = BarChart()
                excel_chart.type = "col"  # Vertikala staplar
                excel_chart.style = 10  # Enkel stil
                excel_chart.barDir = "col"
                excel_chart.grouping = "clustered"
                excel_chart.gapWidth = 150  # Mellanrum mellan staplar

                labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
                data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                excel_chart.add_data(data_ref, titles_from_data=True)
                excel_chart.set_categories(labels)

                # Y-axel - ta bort gridlines för renare utseende
                excel_chart.y_axis.majorGridlines = None
                excel_chart.y_axis.delete = False

                # X-axel
                excel_chart.x_axis.delete = False

                # Datapunktetiketter
                excel_chart.dataLabels = DataLabelList()
                excel_chart.dataLabels.showVal = True
                excel_chart.dataLabels.showCatName = False
                excel_chart.dataLabels.showSerName = False

            # Gemensamma inställningar - Goldman Sachs stil
            excel_chart.title = None  # Titel finns redan ovanför i cellen
            excel_chart.legend = None  # Ingen legend för enkla grafer
            excel_chart.width = 14  # Bredare graf
            excel_chart.height = 8

            # Plot area - vit bakgrund utan ram
            excel_chart.plot_area.layout = None

            # Sätt färger på serier EFTER att data lagts till
            if excel_chart.series:
                for s in excel_chart.series:
                    if chart_type == "line":
                        # Linjefärg navy, tjockare linje
                        s.graphicalProperties.line.solidFill = GS_NAVY
                        s.graphicalProperties.line.width = 28575  # 2.25pt
                        s.smooth = False
                        # Markörpunkter
                        s.marker.symbol = "circle"
                        s.marker.size = 7
                        s.marker.graphicalProperties.solidFill = GS_NAVY
                        s.marker.graphicalProperties.line.solidFill = GS_NAVY
                    elif chart_type == "area":
                        # Ytdiagram - navy fyllning med lite transparens
                        s.graphicalProperties.solidFill = GS_NAVY
                        s.graphicalProperties.line.solidFill = GS_NAVY
                        s.graphicalProperties.line.width = 12700  # 1pt linje
                    elif chart_type != "pie":
                        # Staplar - solid navy fyllning
                        s.graphicalProperties.solidFill = GS_NAVY
                        s.graphicalProperties.line.noFill = True

            # Placera grafen till höger om datan (kolumn D)
            ws.add_chart(excel_chart, f"D{data_start_row - 2}")

        current_row += 2  # Mellanrum mellan grafer

        # Extra mellanrum för att inte grafer ska överlappa
        if data_points:
            current_row += 10

        chart_count += 1

    # Kolumnbredder
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 2  # Mellanrum
    ws.sheet_view.showGridLines = False


def create_separator_sheet(wb, title: str):
    """
    Skapa en separator-flik med en titel.
    Goldman Sachs-stil med navy bakgrund.
    """
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    # Sätt kolumnbredd
    ws.column_dimensions['A'].width = 50

    # Titel i mitten
    ws['A10'] = title.upper()
    ws['A10'].font = Font(name='Arial', size=24, bold=True, color="FFFFFF")
    ws['A10'].alignment = Alignment(horizontal='center', vertical='center')

    # Sätt navy bakgrund på hela arket (via rad-/kolumnformat)
    for row in range(1, 30):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

    return ws


def build_databook(extracted_data: list[dict], output_path: str) -> dict | None:
    """
    Bygg komplett Excel-databok från extraherad data.

    Stödjer både legacy-format (resultatrakning, balansrakning, etc.)
    och nya full-extraktion-formatet (tables med dynamisk struktur).

    Struktur:
    1. [SIFFROR & GRAFER] - separator
    2. Finansiella tabeller (Resultat, Balans, Kassaflöde, etc.)
    3. Grafer
    4. [TEXT] - separator
    5. Textsektioner (VD-kommentar, etc.)

    Args:
        extracted_data: Lista med extraherad data från varje PDF
        output_path: Sökväg för output Excel-fil

    Returns:
        Token-info från AI-normalisering eller None
    """
    if not extracted_data:
        raise ValueError("Ingen data att bygga databok från")

    # Kolla om det är nya formatet (tables) eller legacy (resultatrakning, etc.)
    has_tables = any(d.get("tables") for d in extracted_data)
    has_legacy = any(d.get("resultatrakning") or d.get("balansrakning") for d in extracted_data)

    wb = Workbook()
    wb.remove(wb.active)

    # Sortera data kronologiskt
    sorted_data = sort_by_period(extracted_data)
    periods = [d.get("metadata", {}).get("period", "?") for d in sorted_data]

    # Hämta bolagsnamn
    company_name = sorted_data[0].get("metadata", {}).get("bolag", "Okänt bolag")

    if has_legacy and not has_tables:
        # Legacy-format
        # Separator för siffror
        create_separator_sheet(wb, "═ SIFFROR ═")

        # Flikar för legacy-format
        sheets = [
            ("Resultaträkning", "resultatrakning"),
            ("Balansräkning", "balansrakning"),
            ("Kassaflöde", "kassaflodesanalys"),
        ]

        for sheet_name, data_key in sheets:
            has_data = any(d.get(data_key) for d in sorted_data)
            if has_data:
                ws = wb.create_sheet(sheet_name)
                populate_financial_sheet(ws, sorted_data, data_key, periods, company_name)

    if has_tables:
        # === SEKTION 1: SIFFROR & GRAFER ===
        create_separator_sheet(wb, "═ SIFFROR & GRAFER ═")

        # Nytt format - skapa flikar för varje tabelltyp som finns
        # (map_table_type hanterar quarterly → rätt typ baserat på titel)
        table_types_found = set()
        for item in sorted_data:
            for table in item.get("tables", []):
                table_types_found.add(map_table_type(table))

        # Ordning för flikar
        type_order = ["income_statement", "balance_sheet", "cash_flow", "kpi", "segment", "other"]
        type_sheet_names = {
            "income_statement": "Resultaträkning",
            "balance_sheet": "Balansräkning",
            "cash_flow": "Kassaflöde",
            "kpi": "Nyckeltal",
            "segment": "Segment",
            "other": "Övrigt",
        }

        for table_type in type_order:
            if table_type in table_types_found:
                sheet_name = type_sheet_names.get(table_type, table_type)
                ws = wb.create_sheet(sheet_name)
                populate_dynamic_table_sheet(ws, sorted_data, table_type, company_name)

        # Skapa flik för grafer direkt efter Övrigt
        has_charts = any(item.get("charts") for item in sorted_data)
        if has_charts:
            ws = wb.create_sheet("Grafer")
            populate_charts_sheet(ws, sorted_data, company_name)

        # === SEKTION 2: TEXT ===
        # Samla alla unika sektioner (deduplicera baserat på innehåll)
        def get_bullet_fingerprint(text: str) -> set:
            """Extrahera fingerprint baserat på bullet points (mer distinkt)."""
            # Hitta alla rader som börjar med bullet
            bullets = set()
            for line in text.split('\n'):
                line = line.strip()
                if line.startswith(('• ', '- ', '* ', '– ')):
                    # Ta första 50 tecken efter bullet som fingerprint
                    bullet_text = line[2:52].lower().strip()
                    if bullet_text:
                        bullets.add(bullet_text)
            return bullets

        def is_content_duplicate(c1: str, c2: str) -> bool:
            """Kolla om två sektioner är dubbletter."""
            # Metod 1: Jämför bullet points (bäst för rapporter)
            b1 = get_bullet_fingerprint(c1)
            b2 = get_bullet_fingerprint(c2)
            if b1 and b2:
                # Om minst 3 bullets matchar, är det troligen samma innehåll
                common = b1 & b2
                if len(common) >= 3:
                    return True
                # Om >70% av bullets matchar
                min_bullets = min(len(b1), len(b2))
                if min_bullets > 0 and len(common) / min_bullets > 0.7:
                    return True

            # Metod 2: Exakt matchning av första meningen efter whitespace-normalisering
            def first_sentence(t: str) -> str:
                # Skippa rubriker (korta rader utan punkt)
                for line in t.split('\n'):
                    line = line.strip()
                    if len(line) > 50 and ('.' in line or '•' in line):
                        return line[:100].lower().replace(' ', '')
                return ""

            s1 = first_sentence(c1)
            s2 = first_sentence(c2)
            if s1 and s2 and s1 == s2:
                return True

            return False

        sections_with_page = []
        seen_titles = set()
        seen_contents = []  # Lista med innehåll för likhetsjämförelse

        for item in sorted_data:
            for section in item.get("sections", []):
                title = section.get("title", "")
                content = section.get("content", "")

                # Skippa om vi redan har en sektion med samma innehåll
                is_duplicate = False
                for existing_content in seen_contents:
                    if is_content_duplicate(content, existing_content):
                        is_duplicate = True
                        break

                if is_duplicate:
                    continue

                if title and title not in seen_titles:
                    page = section.get("page", 999)
                    sections_with_page.append((page, title))
                    seen_titles.add(title)
                    seen_contents.append(content)

        # Sortera efter sidnummer (kronologisk ordning)
        sections_with_page.sort(key=lambda x: x[0])

        if sections_with_page:
            create_separator_sheet(wb, "═ TEXT ═")

            for page, section_title in sections_with_page:
                # Sanera fliknamn (tar bort ogiltiga tecken och kortar till 31)
                sheet_name = sanitize_sheet_name(section_title)
                # Undvik duplicerade bladnamn
                existing_sheets = [ws.title for ws in wb.worksheets]
                if sheet_name not in existing_sheets:
                    ws = wb.create_sheet(sheet_name)
                    populate_sections_sheet(ws, sorted_data, section_title, company_name)

    # Spara
    wb.save(output_path)

    return None  # Ingen normalisering längre
